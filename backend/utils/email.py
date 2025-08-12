from fastapi import Depends, HTTPException
from models.mails import Mail
from fastapi_mail import MessageSchema
from models.email_config import fm
import asyncio
from fastapi_mail import MessageSchema
from typing import Union, List



# Updated utils/email.py - Add mail logging
from sqlalchemy.orm import Session
from datetime import time
from crud.mails import create_mail_record_simple
from database import get_db
from crud import storeorders as storeorders_crud
from models.locations import Store
from collections import defaultdict

from database import SessionLocal
import pandas as pd
import os
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


async def send_account_email(email: str, username: str, password: str):
    
    frontend_url = os.getenv("SERVER_URL", "FRONTEND_URL")
    login_link = f"{frontend_url}/signin"    
    
    subject = "Welcome to KPI360.ai - Your Account Credentials"
    
    # HTML email template with proper UI/UX
    html_body = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Welcome to KPI360.ai</title>
        <style>
            body {{
                margin: 0;
                padding: 0;
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background-color: #f5f7fa;
                line-height: 1.6;
            }}
            .email-container {{
                max-width: 600px;
                margin: 0 auto;
                background-color: #ffffff;
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
            }}
            .header {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 40px 30px;
                text-align: center;
                color: white;
            }}
            .logo {{
                font-size: 28px;
                font-weight: bold;
                margin-bottom: 10px;
            }}
            .tagline {{
                font-size: 14px;
                opacity: 0.9;
            }}
            .content {{
                padding: 40px 30px;
            }}
            .welcome-message {{
                font-size: 24px;
                color: #2d3748;
                margin-bottom: 20px;
                font-weight: 600;
            }}
            .message-text {{
                font-size: 16px;
                color: #4a5568;
                margin-bottom: 30px;
            }}
            .credentials-box {{
                background-color: #f7fafc;
                border: 2px solid #e2e8f0;
                border-radius: 8px;
                padding: 25px;
                margin: 25px 0;
            }}
            .credentials-title {{
                font-size: 18px;
                font-weight: 600;
                color: #2d3748;
                margin-bottom: 15px;
                display: flex;
                align-items: center;
            }}
            .credentials-title::before {{
                content: "🔐";
                margin-right: 8px;
            }}
            .credential-item {{
                margin: 12px 0;
                font-size: 14px;
            }}
            .credential-label {{
                font-weight: 600;
                color: #4a5568;
                display: inline-block;
                width: 80px;
            }}
            .credential-value {{
                background-color: #edf2f7;
                padding: 8px 12px;
                border-radius: 4px;
                font-family: 'Courier New', monospace;
                font-size: 14px;
                color: #2d3748;
                border: 1px solid #cbd5e0;
            }}
            .warning-box {{
                background-color: #fef5e7;
                border-left: 4px solid #f6ad55;
                padding: 15px;
                margin: 25px 0;
                border-radius: 4px;
            }}
            .warning-text {{
                color: #744210;
                font-size: 14px;
                margin: 0;
            }}
            .cta-button {{
                display: inline-block;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 14px 30px;
                text-decoration: none;
                border-radius: 6px;
                font-weight: 600;
                margin: 20px 0;
                transition: transform 0.2s ease;
            }}
            .cta-button:hover {{
                transform: translateY(-2px);
            }}
            .footer {{
                background-color: #f7fafc;
                padding: 30px;
                text-align: center;
                border-top: 1px solid #e2e8f0;
            }}
            .footer-text {{
                color: #718096;
                font-size: 14px;
                margin: 5px 0;
            }}
            .social-links {{
                margin-top: 20px;
            }}
            .social-links a {{
                color: #667eea;
                text-decoration: none;
                margin: 0 10px;
                font-size: 14px;
            }}
            @media only screen and (max-width: 600px) {{
                .email-container {{
                    width: 100% !important;
                }}
                .header, .content, .footer {{
                    padding: 20px !important;
                }}
                .welcome-message {{
                    font-size: 20px !important;
                }}
            }}
        </style>
    </head>
    <body>
        <div class="email-container">
            <!-- Header -->
            <div class="header">
                <div class="logo">KPI360.ai</div>
                <div class="tagline">Your Business Intelligence Platform</div>
            </div>
            
            <!-- Content -->
            <div class="content">
                <div class="welcome-message">Welcome aboard, {username}! 🎉</div>
                
                <div class="message-text">
                    Your account has been successfully created for KPI360.ai. We're excited to have you join our platform for advanced business analytics and insights.
                </div>
                
                <!-- Credentials Box -->
                <div class="credentials-box">
                    <div class="credentials-title">Your Login Credentials</div>
                    <div class="credential-item">
                        <span class="credential-label">Email:</span>
                        <div class="credential-value">{email}</div>
                    </div>
                    <div class="credential-item">
                        <span class="credential-label">Password:</span>
                        <div class="credential-value">{password}</div>
                    </div>
                </div>
                
                <!-- Security Warning -->
                <div class="warning-box">
                    <p class="warning-text">
                        <strong>🔒 Important Security Notice:</strong> For your account security, please log in and change your password immediately after your first login.
                    </p>
                </div>
                
                <!-- Call to Action -->
                <div style="text-align: center;">
                    <a href="{login_link}" class="cta-button">Log In to Your Account</a>

                </div>
                
                <div class="message-text">
                    If you have any questions or need assistance, our support team is here to help. Simply reply to this email or contact us through our help center.
                </div>
            </div>
            
            <!-- Footer -->
            <div class="footer">
                <div class="footer-text"><strong>XcelBot Team</strong></div>
                <div class="footer-text">Making data-driven decisions simple and powerful</div>
               
            </div>
        </div>
    </body>
    </html>
    """
    
    # Plain text fallback for email clients that don't support HTML
    text_body = f"""
    Welcome to KPI360.ai, {username}!
    
    Your account has been successfully created.
    
    Login Credentials:
    Email: {email}
    Password: {password}
    
    IMPORTANT: Please log in and change your password as soon as possible for security.
    
    If you need help, contact our support team.
    
    Best regards,
    XcelBot Team
    """
    
    message = MessageSchema(
        subject=subject,
        recipients=[email],
        body=html_body,  # Set HTML as main body
        subtype="html"   # Ensure HTML subtype
    )
    
    await fm.send_message(message)
    
    

async def send_order_confirmation_email(email: str, username: str, order_id: int, items_ordered: dict, order_date, is_update: bool = False):
    subject = "Order Update - KPI360.ai" if is_update else "Order Confirmation - KPI360.ai"
    
    print("Sending email with items:", items_ordered, "to", email, "for user", username, "with order ID", order_id)

    # Format the order date
    formatted_date = order_date.strftime("%B %d, %Y at %I:%M %p")
    
    # Determine the date label and header content based on whether it's an update
    date_label = "Updated At" if is_update else "Order Date"
    header_title = "Order Update" if is_update else "Order Confirmation"
    thank_you_message = f"Your order has been successfully updated, {username}! ✅" if is_update else f"Thank you for your order, {username}! ✅"
    message_text = "Your order has been successfully updated. Here are the details:" if is_update else "Your order has been successfully placed. Here are the details:"
    
    items_list = ""
    for item in items_ordered.get("items", []):
        unit_price = float(item.get('unit_price', 0))
        quantity = int(item.get('quantity', 0))
        total_price = float(item.get('total_price', unit_price * quantity))

        items_list += f"""
        <div style="padding: 10px; border-bottom: 1px solid #e2e8f0; margin-bottom: 8px;">
            <strong>{item.get('name', 'Item')}</strong><br>
            <span style="color: #718096; font-size: 14px;">
                Quantity: {quantity} | 
                Price: ${unit_price:.2f} | 
                Total: ${total_price:.2f}
            </span>
        </div>
        """

    total_amount = sum(float(item.get('total_price', float(item.get('unit_price', 0)) * int(item.get('quantity', 0)))) 
                   for item in items_ordered.get("items", []))

    # HTML email template
    html_body = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>{subject}</title>
        <style>
            body {{
                margin: 0;
                padding: 0;
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background-color: #f5f7fa;
                line-height: 1.6;
            }}
            .email-container {{
                max-width: 600px;
                margin: 0 auto;
                background-color: #ffffff;
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
            }}
            .header {{
                background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
                padding: 40px 30px;
                text-align: center;
                color: white;
            }}
            .logo {{
                font-size: 28px;
                font-weight: bold;
                margin-bottom: 10px;
            }}
            .tagline {{
                font-size: 14px;
                opacity: 0.9;
            }}
            .content {{
                padding: 40px 30px;
            }}
            .order-title {{
                font-size: 24px;
                color: #2d3748;
                margin-bottom: 20px;
                font-weight: 600;
            }}
            .message-text {{
                font-size: 16px;
                color: #4a5568;
                margin-bottom: 30px;
            }}
            .order-details-box {{
                background-color: #f7fafc;
                border: 2px solid #e2e8f0;
                border-radius: 8px;
                padding: 25px;
                margin: 25px 0;
            }}
            .order-info {{
                margin-bottom: 25px;
                padding-bottom: 20px;
                border-bottom: 1px solid #e2e8f0;
            }}
            .info-row {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                padding: 12px 0;
                border-bottom: 1px solid #f1f5f9;
            }}
            .info-row:last-child {{
                border-bottom: none;
            }}
            .info-label {{
                color: #64748b;
                font-weight: 500;
                font-size: 14px;
            }}
            .info-value {{
                color: #1e293b;
                font-weight: 600;
                font-size: 14px;
            }}
            .items-title {{
                font-size: 18px;
                font-weight: 600;
                color: #2d3748;
                margin-bottom: 15px;
                display: flex;
                align-items: center;
            }}
            .items-title::before {{
                content: "📦";
                margin-right: 8px;
            }}
            .total-box {{
                background-color: #48bb78;
                color: white;
                padding: 15px;
                border-radius: 6px;
                text-align: center;
                margin-top: 20px;
            }}
            .total-text {{
                font-size: 18px;
                font-weight: 600;
            }}
            .footer {{
                background-color: #f7fafc;
                padding: 30px;
                text-align: center;
                border-top: 1px solid #e2e8f0;
            }}
            .footer-text {{
                color: #718096;
                font-size: 14px;
                margin: 5px 0;
            }}
            @media only screen and (max-width: 600px) {{
                .email-container {{
                    width: 100% !important;
                }}
                .header, .content, .footer {{
                    padding: 20px !important;
                }}
                .order-title {{
                    font-size: 20px !important;
                }}
                .info-row {{
                    padding: 10px 0 !important;
                }}
            }}
        </style>
    </head>
    <body>
        <div class="email-container">
            <!-- Header -->
            <div class="header">
                <div class="logo">KPI360.ai</div>
                <div class="tagline">{header_title}</div>
            </div>
            
            <!-- Content -->
            <div class="content">
                <div class="order-title">{thank_you_message}</div>
                
                <div class="message-text">
                    {message_text}
                </div>
                
                <!-- Order Details Box -->
                <div class="order-details-box">
                    <div class="order-info">
                        <div class="info-row">
                            <span class="info-label">Order ID:</span>
                            <span class="info-value">#{order_id}</span>
                        </div>
                        <div class="info-row">
                            <span class="info-label">{date_label}:</span>
                            <span class="info-value">{formatted_date}</span>
                        </div>
                        <div class="info-row">
                            <span class="info-label">Total Items:</span>
                            <span class="info-value">{sum(int(item.get("quantity", 0)) for item in items_ordered.get("items", []))}</span>
                        </div>
                    </div>
                    
                    <div class="items-title">Order Items</div>
                    {items_list}
                    
                    <div class="total-box">
                        <div class="total-text">Total Amount: ${total_amount:.2f}</div>
                    </div>
                </div>
                
                <div class="message-text">
                    {"We'll send you another email with tracking information once your order ships." if not is_update else "Your order has been successfully updated with the new items."} 
                    If you have any questions about your order, please don't hesitate to contact our support team.
                </div>
            </div>
            
            <!-- Footer -->
            <div class="footer">
                <div class="footer-text"><strong>KPI360.ai Team</strong></div>
                <div class="footer-text">Thank you for your business!</div>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Plain text fallback
    text_body = f"""
    {subject}
    
    {thank_you_message.replace('✅', '')}
    
    Order Details:
    Order ID: #{order_id}
    {date_label}: {formatted_date}
    Total Items: {items_ordered.get('total_items', 0)}
    
    Items Ordered:
    """
    
    for item in items_ordered.get("items", []):
        unit_price = float(item.get('unit_price', 0))
        quantity = int(item.get('quantity', 0))
        total_price = float(item.get('total_price', unit_price * quantity))
        text_body += f"- {item.get('name', 'Item')}: {quantity} x ${unit_price:.2f} = ${total_price:.2f}\n"
    
    text_body += f"""
    
    Total Amount: ${total_amount:.2f}
    
    {"" if not is_update else "Your order has been successfully updated with the new items."}
    
    Thank you for your business!
    KPI360.ai Team
    """
    
    message = MessageSchema(
        subject=subject,
        recipients=[email],
        body=html_body,
        subtype="html"
    )
    
    await fm.send_message(message)

    




# def get_consolidated_production(company_id: int, db: Session):
#     try:
#         # current_date = 
#                 # Apply date filtering if startDate and endDate are provided
#         startDate  = endDate = datetime.today().date()

#         storeorders = storeorders_crud.get_storeorders_by_company(db, company_id)
#         if not storeorders:
#             return {"message": "No store orders found for this company", "data": []}
        

        
#         # Get all unique locations
#         location_ids = list(set(order.location_id for order in storeorders if order.location_id))
#         locations = db.query(Store).filter(Store.id.in_(location_ids)).all()
#         location_lookup = {loc.id: loc.name for loc in locations}

#         # Build item-location-quantity matrix
#         item_data = defaultdict(lambda: defaultdict(int))  # item_data[item_name][location_name] = quantity
#         item_units = {}  # item_data[item_name] = unit

#         for order in storeorders:
#             location_name = location_lookup.get(order.location_id, f"Location {order.location_id}")
#             if not order.items_ordered or "items" not in order.items_ordered:
#                 continue

#             for item in order.items_ordered["items"]:
#                 name = item.get("name")
#                 quantity = item.get("quantity", 0)
#                 unit = item.get("unit", "")
#                 item_data[name][location_name] += quantity
#                 item_units[name] = unit

#         # Format final table rows
#         all_location_names = sorted(location_lookup.values())
#         table_rows = []
#         for item_name, location_qtys in item_data.items():
#             row = {"Item": item_name}
#             total_required = 0
#             for loc in all_location_names:
#                 qty = location_qtys.get(loc, 0)
#                 row[loc] = qty
#                 total_required += qty
#             row["Total Required"] = total_required
#             row["Unit"] = item_units.get(item_name, "")
#             table_rows.append(row)

#         return {
#             "message": "Consolidated production table generated successfully",
#             "columns": ["Item"] + all_location_names + ["Total Required", "Unit"],
#             "data": table_rows
#         }

#     except Exception as e:
#         raise HTTPException(status_code=500, detail=f"Error fetching consolidated production: {str(e)}")


def get_consolidated_production(company_id: int, db: Session):
    try:
        # Get current date for filtering
        current_date = datetime.today().date()
        
        # Get store orders filtered by current date (prioritizing updated_at over created_at)
        storeorders = storeorders_crud.get_storeorders_by_company_current_date(db, company_id)
        
        # Alternative approach if you prefer to do filtering in this function:
        # all_storeorders = storeorders_crud.get_storeorders_by_company(db, company_id, skip=0, limit=1000)
        # storeorders = []
        # for order in all_storeorders:
        #     order_date = order.updated_at.date() if order.updated_at else order.created_at.date()
        #     if order_date == current_date:
        #         storeorders.append(order)
        
        if not storeorders:
            return {"message": f"No store orders found for this company for {current_date}", "data": []}
        
        # Get all unique locations
        location_ids = list(set(order.location_id for order in storeorders if order.location_id))
        locations = db.query(Store).filter(Store.id.in_(location_ids)).all()
        location_lookup = {loc.id: loc.name for loc in locations}

        # Build item-location-quantity matrix
        item_data = defaultdict(lambda: defaultdict(int))  # item_data[item_name][location_name] = quantity
        item_units = {}  # item_data[item_name] = unit

        for order in storeorders:
            location_name = location_lookup.get(order.location_id, f"Location {order.location_id}")
            if not order.items_ordered or "items" not in order.items_ordered:
                continue

            for item in order.items_ordered["items"]:
                name = item.get("name")
                quantity = item.get("quantity", 0)
                unit = item.get("unit", "")
                if name:  # Only process if name exists
                    item_data[name][location_name] += quantity
                    item_units[name] = unit

        # Format final table rows
        all_location_names = sorted(location_lookup.values())
        table_rows = []
        for item_name, location_qtys in item_data.items():
            row = {"Item": item_name}
            total_required = 0
            for loc in all_location_names:
                qty = location_qtys.get(loc, 0)
                row[loc] = qty
                total_required += qty
            row["Total Required"] = total_required
            row["Unit"] = item_units.get(item_name, "")
            table_rows.append(row)

        return {
            "message": f"Consolidated production table generated successfully for {current_date}",
            "columns": ["Item"] + all_location_names + ["Total Required", "Unit"],
            "data": table_rows
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching consolidated production: {str(e)}")


# from datetime import datetime

# def get_consolidated_production(company_id: int, db: Session):
#     try:
#         # Set date filter to today's date
#         startDate = endDate = datetime.today().date()
        
#         # Get all store orders for the company first
#         all_storeorders = storeorders_crud.get_storeorders_by_company(db, company_id)
#         if not all_storeorders:
#             return {"message": "No store orders found for this company", "data": []}
        
#         # Apply date filtering
#         print(f"Filtering store orders for date: {startDate}")
#         storeorders = []
#         for order in all_storeorders:
#             # Use updated_at if available, otherwise fall back to created_at
#             order_date = None
#             if hasattr(order, 'updated_at') and order.updated_at:
#                 order_date = order.updated_at.date() if hasattr(order.updated_at, 'date') else order.updated_at
#             elif hasattr(order, 'created_at') and order.created_at:
#                 order_date = order.created_at.date() if hasattr(order.created_at, 'date') else order.created_at
            
#             # Filter by date range
#             if order_date and startDate <= order_date <= endDate:
#                 storeorders.append(order)
        
#         print(f"Found {len(storeorders)} orders matching the date filter")
        
#         if not storeorders:
#             return {"message": f"No store orders found for date {startDate}", "data": []}
        
#         # Get all unique locations
#         location_ids = list(set(order.location_id for order in storeorders if order.location_id))
#         locations = db.query(Store).filter(Store.id.in_(location_ids)).all()
#         location_lookup = {loc.id: loc.name for loc in locations}

#         # Build item-location-quantity matrix
#         item_data = defaultdict(lambda: defaultdict(int))  # item_data[item_name][location_name] = quantity
#         item_units = {}  # item_data[item_name] = unit

#         for order in storeorders:
#             location_name = location_lookup.get(order.location_id, f"Location {order.location_id}")
#             if not order.items_ordered or "items" not in order.items_ordered:
#                 continue

#             for item in order.items_ordered["items"]:
#                 name = item.get("name")
#                 quantity = item.get("quantity", 0)
#                 unit = item.get("unit", "")
#                 item_data[name][location_name] += quantity
#                 item_units[name] = unit

#         # Format final table rows
#         all_location_names = sorted(location_lookup.values())
#         table_rows = []
#         for item_name, location_qtys in item_data.items():
#             row = {"Item": item_name}
#             total_required = 0
#             for loc in all_location_names:
#                 qty = location_qtys.get(loc, 0)
#                 row[loc] = qty
#                 total_required += qty
#             row["Total Required"] = total_required
#             row["Unit"] = item_units.get(item_name, "")
#             table_rows.append(row)

#         return {
#             "message": f"Consolidated production table generated successfully for {startDate}",
#             "columns": ["Item"] + all_location_names + ["Total Required", "Unit"],
#             "data": table_rows
#         }

#     except Exception as e:
#         raise HTTPException(status_code=500, detail=f"Error fetching consolidated production: {str(e)}")



# def send_actual_email(to: str, name: str, company_id: int = None):
#     db = SessionLocal()
#     try:
#         # company_id = 1
#         print("---------------------------------------------")
#         print("Sending email to:", to, "for user:", name, "with company ID:", company_id)
#         print("---------------------------------------------")
        
#         data = get_consolidated_production(company_id, db)
        
        
#         print("This is the data which I want to print:", data)

#         # Generate timestamp for file names
#         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
#         # Create downloads directory if it doesn't exist
#         downloads_dir = "downloads"
#         if not os.path.exists(downloads_dir):
#             os.makedirs(downloads_dir)

#         # Generate Excel file
#         def create_excel_file(data, filename):
#             if not data or 'data' not in data or not data['data']:
#                 return None
            
#             try:
#                 # Convert data to DataFrame
#                 df = pd.DataFrame(data['data'])
                
#                 # Create Excel file with formatting
#                 excel_path = os.path.join(downloads_dir, filename)
#                 with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
#                     df.to_excel(writer, sheet_name='Production Requirements', index=False)
                    
#                     # Get workbook and worksheet objects
#                     workbook = writer.book
#                     worksheet = writer.sheets['Production Requirements']
                    
#                     # Apply formatting
#                     from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
#                     # Header formatting
#                     header_font = Font(bold=True, color="000000")
#                     header_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
#                     total_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
#                     highlight_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    
#                     # Border style
#                     thin_border = Border(
#                         left=Side(style='thin'),
#                         right=Side(style='thin'),
#                         top=Side(style='thin'),
#                         bottom=Side(style='thin')
#                     )
                    
#                     # Format headers
#                     for col_num, column in enumerate(df.columns, 1):
#                         cell = worksheet.cell(row=1, column=col_num)
#                         cell.font = header_font
#                         cell.fill = header_fill
#                         cell.alignment = Alignment(horizontal="center", vertical="center")
#                         cell.border = thin_border
                        
#                         # Special formatting for total columns
#                         if 'total' in column.lower() or 'required' in column.lower():
#                             cell.fill = total_fill
                    
#                     # Format data cells
#                     for row_num in range(2, len(df) + 2):
#                         for col_num, column in enumerate(df.columns, 1):
#                             cell = worksheet.cell(row=row_num, column=col_num)
#                             cell.border = thin_border
#                             cell.alignment = Alignment(horizontal="center", vertical="center")
                            
#                             # Highlight cells with values > 0 (except total columns)
#                             if ('total' not in column.lower() and 'required' not in column.lower() 
#                                 and isinstance(cell.value, (int, float)) and cell.value > 0):
#                                 cell.fill = highlight_fill
#                             elif 'total' in column.lower() or 'required' in column.lower():
#                                 cell.fill = total_fill
#                                 cell.font = Font(bold=True)
                    
#                     # Auto-adjust column widths
#                     for column in worksheet.columns:
#                         max_length = 0
#                         column_letter = column[0].column_letter
#                         for cell in column:
#                             try:
#                                 if len(str(cell.value)) > max_length:
#                                     max_length = len(str(cell.value))
#                             except:
#                                 pass
#                         adjusted_width = min(max_length + 2, 50)
#                         worksheet.column_dimensions[column_letter].width = adjusted_width
                
#                 return excel_path
#             except Exception as e:
#                 print(f"Error creating Excel file: {e}")
#                 return None

#         # Generate PDF file
#         def create_pdf_file(data, filename):
#             if not data or 'data' not in data or not data['data']:
#                 return None
            
#             try:
#                 pdf_path = os.path.join(downloads_dir, filename)
                
#                 # Create PDF document
#                 doc = SimpleDocTemplate(pdf_path, pagesize=A4)
#                 elements = []
                
#                 # Get styles
#                 styles = getSampleStyleSheet()
#                 title_style = ParagraphStyle(
#                     'CustomTitle',
#                     parent=styles['Heading1'],
#                     fontSize=16,
#                     spaceAfter=30,
#                     alignment=1  # Center alignment
#                 )
                
#                 # Add title
#                 title = Paragraph("Consolidated Production Requirements", title_style)
#                 elements.append(title)
#                 elements.append(Spacer(1, 12))
                
#                 # Add subtitle
#                 subtitle = Paragraph("", styles['Normal'])
#                 elements.append(subtitle)
#                 elements.append(Spacer(1, 20))
                
#                 # Prepare table data
#                 columns = data.get('columns', list(data['data'][0].keys()) if data['data'] else [])
#                 table_data = [columns]  # Header row
                
#                 # Add data rows
#                 for item in data['data']:
#                     row = [str(item.get(col, "")) for col in columns]
#                     table_data.append(row)
                
#                 # Create table
#                 table = Table(table_data)
                
#                 # Define table style
#                 table_style = [
#                     # Header row styling
#                     ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
#                     ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
#                     ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#                     ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#                     ('FONTSIZE', (0, 0), (-1, 0), 10),
#                     ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    
#                     # Data rows styling
#                     ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
#                     ('FONTSIZE', (0, 1), (-1, -1), 9),
#                     ('GRID', (0, 0), (-1, -1), 1, colors.black),
#                     ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#                 ]
                
#                 # Highlight total columns and positive values
#                 for col_idx, col in enumerate(columns):
#                     if 'total' in col.lower() or 'required' in col.lower():
#                         # Green background for total columns
#                         table_style.append(('BACKGROUND', (col_idx, 0), (col_idx, -1), colors.lightgreen))
#                         table_style.append(('FONTNAME', (col_idx, 1), (col_idx, -1), 'Helvetica-Bold'))
                
#                 # Apply alternating row colors
#                 for row_idx in range(1, len(table_data)):
#                     if row_idx % 2 == 0:
#                         table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.beige))
                
#                 table.setStyle(TableStyle(table_style))
#                 elements.append(table)
                
#                 # Add footer
#                 elements.append(Spacer(1, 30))
#                 footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>KPI360.ai Team"
#                 footer = Paragraph(footer_text, styles['Normal'])
#                 elements.append(footer)
                
#                 # Build PDF
#                 doc.build(elements)
#                 return pdf_path
#             except Exception as e:
#                 print(f"Error creating PDF file: {e}")
#                 return None

#         # Generate HTML table from the data - DYNAMIC VERSION
#         def generate_production_table(data):
#             if not data or 'data' not in data or not data['data']:
#                 return "<p>No production data available.</p>"
            
#             # Get columns dynamically from the data
#             columns = data.get('columns', [])
#             if not columns and data['data']:
#                 # If columns not provided, extract from first row
#                 columns = list(data['data'][0].keys())
            
#             if not columns:
#                 return "<p>No columns found in data.</p>"
            
#             # Start building the table
#             table_html = """
#             <table style="border-collapse: collapse; width: 100%; margin: 20px 0;">
#                 <thead>
#                     <tr style="background-color: #f8f9fa;">
#             """
            
#             # Generate header row dynamically
#             for col in columns:
#                 # Special styling for specific column types
#                 if 'total' in col.lower() or 'required' in col.lower() or 'sum' in col.lower():
#                     header_style = "border: 1px solid #ddd; padding: 12px; text-align: center; font-weight: bold; background-color: #e8f5e8;"
#                 else:
#                     header_style = "border: 1px solid #ddd; padding: 12px; text-align: center; font-weight: bold;"
                
#                 table_html += f'<th style="{header_style}">{col}</th>'
            
#             table_html += """
#                     </tr>
#                 </thead>
#                 <tbody>
#             """
            
#             # Generate data rows dynamically
#             for row_index, item in enumerate(data['data']):
#                 # Alternate row colors for better readability
#                 row_bg = "#f9f9f9" if row_index % 2 == 0 else "#ffffff"
                
#                 table_html += f'<tr style="background-color: {row_bg};">'
                
#                 for col_index, col in enumerate(columns):
#                     cell_value = item.get(col, "")
                    
#                     # Smart styling based on column type and content
#                     if 'total' in col.lower() or 'required' in col.lower() or 'sum' in col.lower():
#                         # Highlight total/required columns in green
#                         cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center; background-color: #e8f5e8; font-weight: bold;"
#                     elif col_index == 0:  # First column (usually item name)
#                         # Check if any store column has a value for this row to highlight the item
#                         has_store_value = any(
#                             isinstance(item.get(c, 0), (int, float)) and item.get(c, 0) > 0 
#                             for c in columns[1:-2] if 'store' in c.lower() or 'location' in c.lower() or 'shop' in c.lower()
#                         )
#                         if has_store_value:
#                             cell_style = "border: 1px solid #ddd; padding: 12px; text-align: left; background-color: #fff3cd;"
#                         else:
#                             cell_style = "border: 1px solid #ddd; padding: 12px; text-align: left;"
#                     elif isinstance(cell_value, (int, float)) and cell_value > 0:
#                         # Highlight cells with positive values (like store quantities)
#                         cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center; background-color: #fff3cd;"
#                     elif col.lower() in ['unit', 'units', 'type', 'category']:
#                         # Unit/type columns - center aligned, no highlighting
#                         cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center;"
#                     else:
#                         # Default styling
#                         cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center;"
                    
#                     table_html += f'<td style="{cell_style}">{cell_value}</td>'
                
#                 table_html += "</tr>"
            
#             table_html += """
#                 </tbody>
#             </table>
#             """
            
#             return table_html

#         # Generate files
#         excel_filename = f"production_requirements_{timestamp}.xlsx"
#         pdf_filename = f"production_requirements_{timestamp}.pdf"
        
#         excel_path = create_excel_file(data, excel_filename)
#         pdf_path = create_pdf_file(data, pdf_filename)
        
#         production_table = generate_production_table(data)
        
#         # Files are attached to email, no download buttons needed in UI
        
#         subject = "Consolidated Production Requirements"
#         html_body = f"""
#         <html>
#         <head>
#             <style>
#                 body {{
#                     font-family: Arial, sans-serif;
#                     line-height: 1.6;
#                     color: #333;
#                     max-width: 1200px;
#                     margin: 0 auto;
#                     padding: 20px;
#                 }}
#                 .header {{
#                     background-color: #f8f9fa;
#                     padding: 20px;
#                     border-radius: 5px;
#                     margin-bottom: 20px;
#                     text-align: center;
#                 }}
#                 .footer {{
#                     margin-top: 30px;
#                     padding-top: 20px;
#                     border-top: 1px solid #ddd;
#                 }}
#                 .table-container {{
#                     overflow-x: auto;
#                 }}
#                 @media screen and (max-width: 600px) {{
#                     .table-container {{
#                         font-size: 12px;
#                     }}
#                 }}
#             </style>
#         </head>
#         <body>
#             <div class="header">
#                 <h2>Consolidated Production Requirements</h2>
#             </div>
            
#             <h3>Hello {name},</h3>
#             <p>Please find below the consolidated production requirements for all stores. Excel and PDF versions are attached to this email for your convenience.</p>
            
#             <div class="table-container">
#                 {production_table}
#             </div>
            
#             <div class="footer">
#                 <p>This report shows the total quantities needed for production across all stores.</p>
#                 <p><strong>Legend:</strong></p>
#                 <ul>
#                     <li>🟡 <strong>Yellow highlighted cells:</strong> Items required by specific stores</li>
#                     <li>🟢 <strong>Green highlighted columns:</strong> Total/Required quantities</li>
#                     <li>📊 <strong>Alternating row colors:</strong> For better readability</li>
#                 </ul>
#                 <p>This report automatically adapts to your data structure.</p>
                
#                 <p>Regards,<br><strong>KPI360.ai Team</strong></p>
#             </div>
#         </body>
#         </html>
#         """

#         # Create attachments list for your MessageSchema
#         attachments = []
        
#         # Add Excel attachment
#         if excel_path and os.path.exists(excel_path):
#             attachments.append(excel_path)
        
#         # Add PDF attachment  
#         if pdf_path and os.path.exists(pdf_path):
#             attachments.append(pdf_path)

#         # Use your original MessageSchema with attachments
#         message = MessageSchema(
#             subject=subject,
#             recipients=[to],
#             body=html_body,
#             subtype="html",
#             attachments=attachments if attachments else None  # Add attachments if they exist
#         )

#         async def send():
#             await fm.send_message(message)

#         asyncio.run(send())
#         print(f"Production requirements email with downloads sent successfully to {to}")
        
#         # Clean up files after sending
#         if excel_path and os.path.exists(excel_path):
#             os.remove(excel_path)
#         if pdf_path and os.path.exists(pdf_path):
#             os.remove(pdf_path)
        
#     except Exception as e:
#         print(f"Failed to send production requirements email to {to}: {e}")
#     finally:
#         db.close()
        
        

def send_actual_email(to: str, name: str, company_id: int = None):
    db = SessionLocal()
    try:
        # company_id = 1
        print("---------------------------------------------")
        print("Sending email to:", to, "for user:", name, "with company ID:", company_id)
        print("---------------------------------------------")
        
        data = get_consolidated_production(company_id, db)
        
        # Custom location ordering function
        def reorder_data_columns(data):
            """Reorder the data columns according to the specified location priority"""
            if not data or 'data' not in data or not data['data']:
                return data
            
            # Define the priority order (case-insensitive matching)
            priority_locations = [
                # Priority 1: Midtown East variations
                ["midtown east", "Midtown East"],
                # Priority 2: Lenox Hill variations  
                ["lenox hill", "Lenox Hill", "lenox hills", "Lenox Hills"],
                # Priority 3: Hell's Kitchen variations
                ["hells kitchen", "Hells Kitchen", "hell's kitchen", "Hell's Kitchen"],
                # Priority 4: Union Square variations
                ["union square", "Union Square", "Union square"],
                # Priority 5: Flatiron variations
                ["flatiron", "Flatiron", "FLATIRON"],
                # Priority 6: Williamsburg variations
                ["williamsburg", "Williamsburg"]
            ]
            
            # Get original columns
            original_columns = data.get('columns', [])
            if not original_columns:
                return data
            
            # Separate fixed columns (Item, Total Required, Unit) from location columns
            fixed_start = ["Item"]
            fixed_end = ["Total Required", "Unit"]
            
            # Find location columns (everything except fixed columns)
            location_columns = []
            for col in original_columns:
                if col not in fixed_start + fixed_end:
                    location_columns.append(col)
            
            # Function to find matching priority group for a location
            def get_priority_group(location_name):
                location_lower = location_name.lower().strip()
                for i, priority_group in enumerate(priority_locations):
                    for variant in priority_group:
                        if variant.lower().strip() == location_lower:
                            return i
                return len(priority_locations)  # Return high number for non-priority locations
            
            # Sort locations by priority
            sorted_locations = sorted(location_columns, key=lambda x: (get_priority_group(x), x.lower()))
            
            # Create new column order
            new_columns = fixed_start + sorted_locations + fixed_end
            
            # Reorder data rows to match new column order
            reordered_data = []
            for row in data['data']:
                new_row = {}
                for col in new_columns:
                    new_row[col] = row.get(col, "")
                reordered_data.append(new_row)
            
            return {
                "message": data.get('message', ''),
                "columns": new_columns,
                "data": reordered_data
            }
        
        # Reorder the data according to custom location priority
        data = reorder_data_columns(data)
        
        print("This is the data which I want to print:", data)

        # Generate timestamp for file names
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create downloads directory if it doesn't exist
        downloads_dir = "downloads"
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)

        # Generate Excel file
        def create_excel_file(data, filename):
            if not data or 'data' not in data or not data['data']:
                return None
            
            try:
                # Convert data to DataFrame - now with reordered columns
                df = pd.DataFrame(data['data'])
                # Ensure column order matches our custom order
                if 'columns' in data:
                    df = df[data['columns']]
                
                # Create Excel file with formatting
                excel_path = os.path.join(downloads_dir, filename)
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Production Requirements', index=False)
                    
                    # Get workbook and worksheet objects
                    workbook = writer.book
                    worksheet = writer.sheets['Production Requirements']
                    
                    # Apply formatting
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    # Header formatting
                    header_font = Font(bold=True, color="000000")
                    header_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                    total_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    highlight_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    
                    # Border style
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Format headers
                    for col_num, column in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                        
                        # Special formatting for total columns
                        if 'total' in column.lower() or 'required' in column.lower():
                            cell.fill = total_fill
                    
                    # Format data cells
                    for row_num in range(2, len(df) + 2):
                        for col_num, column in enumerate(df.columns, 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            # Highlight cells with values > 0 (except total columns)
                            if ('total' not in column.lower() and 'required' not in column.lower() 
                                and isinstance(cell.value, (int, float)) and cell.value > 0):
                                cell.fill = highlight_fill
                            elif 'total' in column.lower() or 'required' in column.lower():
                                cell.fill = total_fill
                                cell.font = Font(bold=True)
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                return excel_path
            except Exception as e:
                print(f"Error creating Excel file: {e}")
                return None

        # Generate PDF file
        def create_pdf_file(data, filename):
            if not data or 'data' not in data or not data['data']:
                return None
            
            try:
                pdf_path = os.path.join(downloads_dir, filename)
                
                # Create PDF document
                doc = SimpleDocTemplate(pdf_path, pagesize=A4)
                elements = []
                
                # Get styles
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    spaceAfter=30,
                    alignment=1  # Center alignment
                )
                
                # Add title
                title = Paragraph("Consolidated Production Requirements", title_style)
                elements.append(title)
                elements.append(Spacer(1, 12))
                
                # Add subtitle
                subtitle = Paragraph("", styles['Normal'])
                elements.append(subtitle)
                elements.append(Spacer(1, 20))
                
                # Prepare table data - now with custom column order
                columns = data.get('columns', list(data['data'][0].keys()) if data['data'] else [])
                table_data = [columns]  # Header row
                
                # Add data rows
                for item in data['data']:
                    row = [str(item.get(col, "")) for col in columns]
                    table_data.append(row)
                
                # Create table
                table = Table(table_data)
                
                # Define table style
                table_style = [
                    # Header row styling
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    
                    # Data rows styling
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]
                
                # Highlight total columns and positive values
                for col_idx, col in enumerate(columns):
                    if 'total' in col.lower() or 'required' in col.lower():
                        # Green background for total columns
                        table_style.append(('BACKGROUND', (col_idx, 0), (col_idx, -1), colors.lightgreen))
                        table_style.append(('FONTNAME', (col_idx, 1), (col_idx, -1), 'Helvetica-Bold'))
                
                # Apply alternating row colors
                for row_idx in range(1, len(table_data)):
                    if row_idx % 2 == 0:
                        table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.beige))
                
                table.setStyle(TableStyle(table_style))
                elements.append(table)
                
                # Add footer
                elements.append(Spacer(1, 30))
                footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>KPI360.ai Team"
                footer = Paragraph(footer_text, styles['Normal'])
                elements.append(footer)
                
                # Build PDF
                doc.build(elements)
                return pdf_path
            except Exception as e:
                print(f"Error creating PDF file: {e}")
                return None

        # Generate HTML table from the data - DYNAMIC VERSION with custom ordering
        def generate_production_table(data):
            if not data or 'data' not in data or not data['data']:
                return "<p>No production data available.</p>"
            
            # Get columns dynamically from the reordered data
            columns = data.get('columns', [])
            if not columns and data['data']:
                # If columns not provided, extract from first row
                columns = list(data['data'][0].keys())
            
            if not columns:
                return "<p>No columns found in data.</p>"
            
            # Start building the table
            table_html = """
            <table style="border-collapse: collapse; width: 100%; margin: 20px 0;">
                <thead>
                    <tr style="background-color: #f8f9fa;">
            """
            
            # Generate header row dynamically
            for col in columns:
                # Special styling for specific column types
                if 'total' in col.lower() or 'required' in col.lower() or 'sum' in col.lower():
                    header_style = "border: 1px solid #ddd; padding: 12px; text-align: center; font-weight: bold; background-color: #e8f5e8;"
                else:
                    header_style = "border: 1px solid #ddd; padding: 12px; text-align: center; font-weight: bold;"
                
                table_html += f'<th style="{header_style}">{col}</th>'
            
            table_html += """
                    </tr>
                </thead>
                <tbody>
            """
            
            # Generate data rows dynamically
            for row_index, item in enumerate(data['data']):
                # Alternate row colors for better readability
                row_bg = "#f9f9f9" if row_index % 2 == 0 else "#ffffff"
                
                table_html += f'<tr style="background-color: {row_bg};">'
                
                for col_index, col in enumerate(columns):
                    cell_value = item.get(col, "")
                    
                    # Smart styling based on column type and content
                    if 'total' in col.lower() or 'required' in col.lower() or 'sum' in col.lower():
                        # Highlight total/required columns in green
                        cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center; background-color: #e8f5e8; font-weight: bold;"
                    elif col_index == 0:  # First column (usually item name)
                        # Check if any store column has a value for this row to highlight the item
                        has_store_value = any(
                            isinstance(item.get(c, 0), (int, float)) and item.get(c, 0) > 0 
                            for c in columns[1:-2] if c not in ["Item", "Total Required", "Unit"]
                        )
                        if has_store_value:
                            cell_style = "border: 1px solid #ddd; padding: 12px; text-align: left; background-color: #fff3cd;"
                        else:
                            cell_style = "border: 1px solid #ddd; padding: 12px; text-align: left;"
                    elif isinstance(cell_value, (int, float)) and cell_value > 0:
                        # Highlight cells with positive values (like store quantities)
                        cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center; background-color: #fff3cd;"
                    elif col.lower() in ['unit', 'units', 'type', 'category']:
                        # Unit/type columns - center aligned, no highlighting
                        cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center;"
                    else:
                        # Default styling
                        cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center;"
                    
                    table_html += f'<td style="{cell_style}">{cell_value}</td>'
                
                table_html += "</tr>"
            
            table_html += """
                </tbody>
            </table>
            """
            
            return table_html

        # Generate files
        excel_filename = f"production_requirements_{timestamp}.xlsx"
        pdf_filename = f"production_requirements_{timestamp}.pdf"
        
        excel_path = create_excel_file(data, excel_filename)
        pdf_path = create_pdf_file(data, pdf_filename)
        
        production_table = generate_production_table(data)
        
        # Files are attached to email, no download buttons needed in UI
        
        subject = "Consolidated Production Requirements"
        html_body = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    max-width: 1200px;
                    margin: 0 auto;
                    padding: 20px;
                }}
                .header {{
                    background-color: #f8f9fa;
                    padding: 20px;
                    border-radius: 5px;
                    margin-bottom: 20px;
                    text-align: center;
                }}
                .footer {{
                    margin-top: 30px;
                    padding-top: 20px;
                    border-top: 1px solid #ddd;
                }}
                .table-container {{
                    overflow-x: auto;
                }}
                @media screen and (max-width: 600px) {{
                    .table-container {{
                        font-size: 12px;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>Consolidated Production Requirements</h2>
            </div>
            
            <h3>Hello {name},</h3>
            <p>Please find below the consolidated production requirements for all stores. Excel and PDF versions are attached to this email for your convenience.</p>
            
            <div class="table-container">
                {production_table}
            </div>
            
            <div class="footer">
                <p>This report shows the total quantities needed for production across all stores.</p>
                <p><strong>Legend:</strong></p>
                <ul>
                    <li>🟡 <strong>Yellow highlighted cells:</strong> Items required by specific stores</li>
                    <li>🟢 <strong>Green highlighted columns:</strong> Total/Required quantities</li>
                    <li>📊 <strong>Alternating row colors:</strong> For better readability</li>
                </ul>
                <p>This report automatically adapts to your data structure with prioritized location ordering.</p>
                
                <p>Regards,<br><strong>KPI360.ai Team</strong></p>
            </div>
        </body>
        </html>
        """

        # Create attachments list for your MessageSchema
        attachments = []
        
        # Add Excel attachment
        if excel_path and os.path.exists(excel_path):
            attachments.append(excel_path)
        
        # Add PDF attachment  
        if pdf_path and os.path.exists(pdf_path):
            attachments.append(pdf_path)

        # Use your original MessageSchema with attachments
        message = MessageSchema(
            subject=subject,
            recipients=[to],
            body=html_body,
            subtype="html",
            attachments=attachments if attachments else None  # Add attachments if they exist
        )

        async def send():
            await fm.send_message(message)

        asyncio.run(send())
        print(f"Production requirements email with downloads sent successfully to {to}")
        
        # Clean up files after sending
        if excel_path and os.path.exists(excel_path):
            os.remove(excel_path)
        if pdf_path and os.path.exists(pdf_path):
            os.remove(pdf_path)
        
    except Exception as e:
        print(f"Failed to send production requirements email to {to}: {e}")
    finally:
        db.close()

        
# def send_production(to: Union[str, List[str]], name: str, company_id: int = None, is_update: bool = False, is_email_update: bool = False, recent_order_id: int = None):
#     """
#     Send production requirements email with attachments
    
#     Args:
#         to: Recipient email address (string) or list of email addresses
#         name: Recipient name
#         company_id: Company ID for data filtering
#         is_update: Boolean flag to indicate if this is an update to existing requirements
#         is_email_update: Boolean flag to indicate if items were ordered after global time
#         recent_order_id: ID of the recent order to highlight (when email sent after global time)
#     """
#     db = SessionLocal()
#     try:
#         # Handle both single email and list of emails
#         if isinstance(to, str):
#             recipient_list = [to]
#         elif isinstance(to, list):
#             recipient_list = to
#         else:
#             print(f"Invalid email format: {to}")
#             return
        
#         print("---------------------------------------------")
#         print("Sending production email to:", recipient_list, "for user:", name, "with company ID:", company_id, "Is update:", is_update, "Is email update:", is_email_update, "Recent order ID:", recent_order_id)
#         print("---------------------------------------------")
        
#         # Get global time for highlighting recent orders
#         global_time = None
#         recent_order_items = set()
#         current_order_items = set()  # Initialize here to ensure it's always available
        
#         try:
#             first_company_email = db.query(Mail).filter(Mail.company_id == company_id).first()
#             if first_company_email and first_company_email.receiving_time:
#                 global_time = first_company_email.receiving_time
#                 print(f"Global time for highlighting: {global_time}")
#         except Exception as e:
#             print(f"Error getting global time: {e}")
        
#         data = get_consolidated_production(company_id, db)
#         print("Production data retrieved:", bool(data))
        
#         if not data:
#             print("WARNING: No production data available")
#             return  # Exit early if no data
        
#         # Get recent orders (after global time) for highlighting
#         if global_time:
#             try:
#                 from sqlalchemy import and_
#                 # Import StoreOrders model - adjust import path as needed
#                 from models.storeorders import StoreOrders
                
#                 # Get today's date to combine with global_time
#                 today = datetime.now().date()
#                 global_datetime_today = datetime.combine(today, global_time)
                
#                 # Query orders created after global time
#                 recent_orders = db.query(StoreOrders).filter(
#                     and_(
#                         StoreOrders.company_id == company_id,
#                         StoreOrders.created_at > global_datetime_today
#                     )
#                 ).all()
                
#                 # Extract item names from recent orders
#                 for order in recent_orders:
#                     if order.items_ordered and isinstance(order.items_ordered, dict) and 'items' in order.items_ordered:
#                         for item in order.items_ordered['items']:
#                             if isinstance(item, dict):
#                                 if 'item_name' in item:
#                                     recent_order_items.add(str(item['item_name']))
#                                 elif 'name' in item:
#                                     recent_order_items.add(str(item['name']))
                
#                 # If we have a specific recent_order_id, get its items for special highlighting
#                 if recent_order_id:
#                     current_order = db.query(StoreOrders).filter(StoreOrders.id == recent_order_id).first()
#                     if current_order and current_order.items_ordered and isinstance(current_order.items_ordered, dict) and 'items' in current_order.items_ordered:
#                         for item in current_order.items_ordered['items']:
#                             if isinstance(item, dict):
#                                 if 'item_name' in item:
#                                     current_order_items.add(str(item['item_name']))
#                                 elif 'name' in item:
#                                     current_order_items.add(str(item['name']))
#                         print(f"Current order items to specially highlight: {current_order_items}")
                
#                 print(f"Recent order items to highlight: {recent_order_items}")
#             except Exception as e:
#                 print(f"Error getting recent orders: {e}")
#                 # Continue without highlighting if there's an error
        
#         print("This is the data which I want to print:", data)

#         # Generate timestamp for file names
#         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
#         # Create downloads directory if it doesn't exist
#         downloads_dir = "downloads"
#         if not os.path.exists(downloads_dir):
#             os.makedirs(downloads_dir)

#         # Generate Excel file
#         def create_excel_file(data, filename):
#             if not data or 'data' not in data or not data['data']:
#                 return None
            
#             try:
#                 # Convert data to DataFrame
#                 df = pd.DataFrame(data['data'])
                
#                 # Create Excel file with formatting
#                 excel_path = os.path.join(downloads_dir, filename)
#                 with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
#                     df.to_excel(writer, sheet_name='Production Requirements', index=False)
                    
#                     # Get workbook and worksheet objects
#                     workbook = writer.book
#                     worksheet = writer.sheets['Production Requirements']
                    
#                     # Apply formatting
#                     from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
#                     # Header formatting
#                     header_font = Font(bold=True, color="000000")
#                     header_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
#                     total_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
#                     highlight_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
#                     recent_order_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red for recent orders
#                     current_order_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Darker red for current order
                    
#                     # Border style
#                     thin_border = Border(
#                         left=Side(style='thin'),
#                         right=Side(style='thin'),
#                         top=Side(style='thin'),
#                         bottom=Side(style='thin')
#                     )
                    
#                     # Get item name column index
#                     item_name_col = None
#                     for idx, col in enumerate(df.columns):
#                         if 'item' in col.lower() and ('name' in col.lower() or col.lower() == 'item'):
#                             item_name_col = idx + 1  # Excel is 1-indexed
#                             break
                    
#                     # Format headers
#                     for col_num, column in enumerate(df.columns, 1):
#                         cell = worksheet.cell(row=1, column=col_num)
#                         cell.font = header_font
#                         cell.fill = header_fill
#                         cell.alignment = Alignment(horizontal="center", vertical="center")
#                         cell.border = thin_border
                        
#                         # Special formatting for total columns
#                         if 'total' in column.lower() or 'required' in column.lower():
#                             cell.fill = total_fill
                    
#                     # Format data cells
#                     for row_num in range(2, len(df) + 2):
#                         # Check if this item is a recent order or current order
#                         is_recent_item = False
#                         is_current_order_item = False
#                         if item_name_col and (recent_order_items or current_order_items):
#                             item_name = worksheet.cell(row=row_num, column=item_name_col).value
#                             item_name_str = str(item_name)
#                             is_current_order_item = item_name_str in current_order_items
#                             # Recent items include ALL items ordered after global time (including current order)
#                             is_recent_item = item_name_str in recent_order_items
                        
#                         for col_num, column in enumerate(df.columns, 1):
#                             cell = worksheet.cell(row=row_num, column=col_num)
#                             cell.border = thin_border
#                             cell.alignment = Alignment(horizontal="center", vertical="center")
                            
#                             # Apply base effects first: 1) Recent orders (red), 2) Totals (green), 3) Positive values (yellow)
#                             if is_recent_item:
#                                 cell.fill = recent_order_fill
#                                 cell.font = Font(bold=True, color="CC0000")  # Red bold text
#                             elif 'total' in column.lower() or 'required' in column.lower():
#                                 cell.fill = total_fill
#                                 cell.font = Font(bold=True)
#                             elif ('total' not in column.lower() and 'required' not in column.lower() 
#                                 and isinstance(cell.value, (int, float)) and cell.value > 0):
#                                 cell.fill = highlight_fill
                            
#                             # Final layer: Apply bold border and underline for current order items (keep existing formatting)
#                             if is_current_order_item:
#                                 # Keep existing font but add underline
#                                 current_font = cell.font
#                                 cell.font = Font(
#                                     bold=current_font.bold,
#                                     color=current_font.color,
#                                     italic=False,  # No italic
#                                     underline='single'
#                                 )
#                                 # Add medium border for current order items
#                                 medium_border = Border(
#                                     left=Side(style='medium'),
#                                     right=Side(style='medium'),
#                                     top=Side(style='medium'),
#                                     bottom=Side(style='medium')
#                                 )
#                                 cell.border = medium_border
                    
#                     # Auto-adjust column widths
#                     for column in worksheet.columns:
#                         max_length = 0
#                         column_letter = column[0].column_letter
#                         for cell in column:
#                             try:
#                                 if len(str(cell.value)) > max_length:
#                                     max_length = len(str(cell.value))
#                             except:
#                                 pass
#                         adjusted_width = min(max_length + 2, 50)
#                         worksheet.column_dimensions[column_letter].width = adjusted_width
                
#                 return excel_path
#             except Exception as e:
#                 print(f"Error creating Excel file: {e}")
#                 return None

#         # Generate PDF file
#         def create_pdf_file(data, filename):
#             if not data or 'data' not in data or not data['data']:
#                 return None
            
#             try:
#                 pdf_path = os.path.join(downloads_dir, filename)
                
#                 # Create PDF document
#                 doc = SimpleDocTemplate(pdf_path, pagesize=A4)
#                 elements = []
                
#                 # Get styles
#                 styles = getSampleStyleSheet()
#                 title_style = ParagraphStyle(
#                     'CustomTitle',
#                     parent=styles['Heading1'],
#                     fontSize=16,
#                     spaceAfter=30,
#                     alignment=1  # Center alignment
#                 )
                
#                 # Add title with update indicator if needed
#                 title_text = "Consolidated Production Requirements"
#                 if is_update:
#                     title_text += " (Updated)"
#                 title = Paragraph(title_text, title_style)
#                 elements.append(title)
#                 elements.append(Spacer(1, 12))
                
#                 # Add subtitle
#                 subtitle = Paragraph("", styles['Normal'])
#                 elements.append(subtitle)
#                 elements.append(Spacer(1, 20))
                
#                 # Prepare table data
#                 columns = data.get('columns', list(data['data'][0].keys()) if data['data'] else [])
#                 table_data = [columns]  # Header row
                
#                 # Add data rows
#                 for item in data['data']:
#                     row = [str(item.get(col, "")) for col in columns]
#                     table_data.append(row)
                
#                 # Create table
#                 table = Table(table_data)
                
#                 # Define table style
#                 table_style = [
#                     # Header row styling
#                     ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
#                     ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
#                     ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#                     ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#                     ('FONTSIZE', (0, 0), (-1, 0), 10),
#                     ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    
#                     # Data rows styling
#                     ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
#                     ('FONTSIZE', (0, 1), (-1, -1), 9),
#                     ('GRID', (0, 0), (-1, -1), 1, colors.black),
#                     ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#                 ]
                
#                 # Highlight total columns and positive values
#                 for col_idx, col in enumerate(columns):
#                     if 'total' in col.lower() or 'required' in col.lower():
#                         # Green background for total columns
#                         table_style.append(('BACKGROUND', (col_idx, 0), (col_idx, -1), colors.lightgreen))
#                         table_style.append(('FONTNAME', (col_idx, 1), (col_idx, -1), 'Helvetica-Bold'))
                
#                 # Apply alternating row colors
#                 for row_idx in range(1, len(table_data)):
#                     if row_idx % 2 == 0:
#                         table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.beige))
                
#                 table.setStyle(TableStyle(table_style))
#                 elements.append(table)
                
#                 # Add footer
#                 elements.append(Spacer(1, 30))
#                 footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>KPI360.ai Team"
#                 footer = Paragraph(footer_text, styles['Normal'])
#                 elements.append(footer)
                
#                 # Build PDF
#                 doc.build(elements)
#                 return pdf_path
#             except Exception as e:
#                 print(f"Error creating PDF file: {e}")
#                 return None

#         # Generate HTML table from the data - DYNAMIC VERSION
#         def generate_production_table(data):
#             if not data or 'data' not in data or not data['data']:
#                 return "<p>No production data available.</p>"
            
#             # Get columns dynamically from the data
#             columns = data.get('columns', [])
#             if not columns and data['data']:
#                 # If columns not provided, extract from first row
#                 columns = list(data['data'][0].keys())
            
#             if not columns:
#                 return "<p>No columns found in data.</p>"
            
#             # Find item name column
#             item_name_col = None
#             for idx, col in enumerate(columns):
#                 if 'item' in col.lower() and ('name' in col.lower() or col.lower() == 'item'):
#                     item_name_col = idx
#                     break
            
#             # Start building the table
#             table_html = """
#             <table style="border-collapse: collapse; width: 100%; margin: 20px 0;">
#                 <thead>
#                     <tr style="background-color: #f8f9fa;">
#             """
            
#             # Generate header row dynamically
#             for col in columns:
#                 # Special styling for specific column types
#                 if 'total' in col.lower() or 'required' in col.lower() or 'sum' in col.lower():
#                     header_style = "border: 1px solid #ddd; padding: 12px; text-align: center; font-weight: bold; background-color: #e8f5e8;"
#                 else:
#                     header_style = "border: 1px solid #ddd; padding: 12px; text-align: center; font-weight: bold;"
                
#                 table_html += f'<th style="{header_style}">{col}</th>'
            
#             table_html += """
#                     </tr>
#                 </thead>
#                 <tbody>
#             """
            
#             # Generate data rows dynamically
#             for row_index, item in enumerate(data['data']):
#                 # Check if this item is a recent order or current order
#                 is_recent_item = False
#                 is_current_order_item = False
#                 if item_name_col is not None and (recent_order_items or current_order_items):
#                     item_name = list(item.values())[item_name_col] if item_name_col < len(item) else None
#                     item_name_str = str(item_name)
#                     is_current_order_item = item_name_str in current_order_items
#                     # Recent items include ALL items ordered after global time (including current order)
#                     is_recent_item = item_name_str in recent_order_items
                
#                 # Alternate row colors for better readability, with special highlighting for orders
#                 if is_recent_item:
#                     row_bg = "#ffe6e6"  # Light red for recent orders
#                 else:
#                     row_bg = "#f9f9f9" if row_index % 2 == 0 else "#ffffff"
                
#                 table_html += f'<tr style="background-color: {row_bg};">'
                
#                 for col_index, col in enumerate(columns):
#                     cell_value = item.get(col, "")
                    
#                     # Apply base effects first: 1) Recent orders (red), 2) Totals (green), 3) Positive values (yellow)
#                     if is_recent_item and col_index == item_name_col:
#                         # Recent item name - bold red text
#                         cell_style = "border: 1px solid #ddd; padding: 12px; text-align: left; background-color: #ffe6e6; font-weight: bold; color: #cc0000;"
#                     elif is_recent_item:
#                         # Recent item other columns
#                         cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center; background-color: #ffe6e6; font-weight: bold; color: #cc0000;"
#                     elif 'total' in col.lower() or 'required' in col.lower() or 'sum' in col.lower():
#                         # Highlight total/required columns in green
#                         cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center; background-color: #e8f5e8; font-weight: bold;"
#                     elif col_index == 0:  # First column (usually item name)
#                         # Check if any store column has a value for this row to highlight the item
#                         has_store_value = any(
#                             isinstance(item.get(c, 0), (int, float)) and item.get(c, 0) > 0 
#                             for c in columns[1:-2] if 'store' in c.lower() or 'location' in c.lower() or 'shop' in c.lower()
#                         )
#                         if has_store_value:
#                             cell_style = "border: 1px solid #ddd; padding: 12px; text-align: left; background-color: #fff3cd;"
#                         else:
#                             cell_style = "border: 1px solid #ddd; padding: 12px; text-align: left;"
#                     elif isinstance(cell_value, (int, float)) and cell_value > 0:
#                         # Highlight cells with positive values (like store quantities)
#                         cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center; background-color: #fff3cd;"
#                     elif col.lower() in ['unit', 'units', 'type', 'category']:
#                         # Unit/type columns - center aligned, no highlighting
#                         cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center;"
#                     else:
#                         # Default styling
#                         cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center;"
                    
#                     # Final layer: Apply medium border and underline for current order items (keep base effects)
#                     if is_current_order_item:
#                         # Add underline and medium border to existing styling
#                         cell_style += " text-decoration: underline; border: 2px solid #333;"
                    
#                     table_html += f'<td style="{cell_style}">{cell_value}</td>'
                
#                 table_html += "</tr>"
            
#             table_html += """
#                 </tbody>
#             </table>
#             """
            
#             return table_html

#         # Generate files
#         excel_filename = f"production_requirements_{timestamp}.xlsx"
#         pdf_filename = f"production_requirements_{timestamp}.pdf"
        
#         excel_path = create_excel_file(data, excel_filename)
#         pdf_path = create_pdf_file(data, pdf_filename)
        
#         production_table = generate_production_table(data)
        
#         # Determine subject and greeting based on flags
#         subject = "Consolidated Production Requirements"
#         greeting = f"Hello {name},"
        
#         # Determine intro text based on is_email_update flag
#         if is_email_update:
#             intro_text = "**Updated** - Please find below the *updated* consolidated production requirements for all stores."
#         else:
#             intro_text = "Please find below the consolidated production requirements for all stores."
        
#         # Add "Updated" to subject if is_update is True
#         if is_update:
#             subject += " - Updated"
        
#         html_body = f"""
#         <html>
#         <head>
#             <style>
#                 body {{
#                     font-family: Arial, sans-serif;
#                     line-height: 1.6;
#                     color: #333;
#                     max-width: 1200px;
#                     margin: 0 auto;
#                     padding: 20px;
#                 }}
#                 .header {{
#                     background-color: #f8f9fa;
#                     padding: 20px;
#                     border-radius: 5px;
#                     margin-bottom: 20px;
#                     text-align: center;
#                 }}
#                 .footer {{
#                     margin-top: 30px;
#                     padding-top: 20px;
#                     border-top: 1px solid #ddd;
#                 }}
#                 .table-container {{
#                     overflow-x: auto;
#                 }}
#                 @media screen and (max-width: 600px) {{
#                     .table-container {{
#                         font-size: 12px;
#                     }}
#                 }}
#                 .update-notice {{
#                     background-color: #fff3cd;
#                     border: 1px solid #ffeaa7;
#                     border-radius: 5px;
#                     padding: 15px;
#                     margin-bottom: 20px;
#                 }}
#             </style>
#         </head>
#         <body>
#             <div class="header">
#                 <h2>{subject}</h2>
#             </div>
            
#             <h3>{greeting}</h3>
            
#             {"<div class='update-notice'><strong>📋 Update Notice:</strong> This is an updated version of the production requirements.</div>" if is_update else ""}
            
#             {"<div class='update-notice'><strong>🔔 New Orders Alert:</strong> Some items were ordered after the global time and are highlighted in red." + (f" Items from Order ID #{recent_order_id} also have medium borders and underlines." if recent_order_id else "") + "</div>" if is_email_update else ""}
            
#             <p>{intro_text} Excel and PDF versions are attached to this email for your convenience.</p>
            
#             <div class="table-container">
#                 {production_table}
#             </div>
            
#             <div class="footer">
#                 <p>This report shows the total quantities needed for production across all stores.</p>
#                 <p><strong>Legend:</strong></p>
#                 <ul>
#                     <li>🔴 <strong>Red highlighted rows:</strong> Items ordered after global time (recent orders)</li>
#                     <li>🟡 <strong>Yellow highlighted cells:</strong> Items required by specific stores</li>
#                     <li>🟢 <strong>Green highlighted columns:</strong> Total/Required quantities</li>
#                     <li>📊 <strong>Alternating row colors:</strong> For better readability</li>
#                     {f"<li>🔲 <strong><u>Medium borders and underlined:</u></strong> Items from current order (Order #{recent_order_id})</li>" if recent_order_id else ""}
#                 </ul>
#                 <p>This report automatically adapts to your data structure.</p>
                
#                 <p>Regards,<br><strong>KPI360.ai Team</strong></p>
#             </div>
#         </body>
#         </html>
#         """

#         # Create attachments list for your MessageSchema
#         attachments = []
        
#         # Add Excel attachment
#         if excel_path and os.path.exists(excel_path):
#             attachments.append(excel_path)
        
#         # Add PDF attachment  
#         if pdf_path and os.path.exists(pdf_path):
#             attachments.append(pdf_path)

#         # Use your original MessageSchema with attachments
#         message = MessageSchema(
#             subject=subject,
#             recipients=recipient_list,  # Now supports multiple recipients
#             body=html_body,
#             subtype="html",
#             attachments=attachments if attachments else None  # Add attachments if they exist
#         )

#         async def send():
#             try:
#                 await fm.send_message(message)
#                 print(f"Email sent successfully to {recipient_list}")
#             except Exception as e:
#                 print(f"Error sending email: {e}")
#                 raise e

#         try:
#             asyncio.run(send())
#             print(f"Production requirements email {'(update)' if is_update else ''} {'(with recent orders)' if is_email_update else ''} with downloads sent successfully to {recipient_list}")
#         except Exception as e:
#             print(f"Failed to send email via asyncio: {e}")
#             # Try alternative approach if asyncio.run fails
#             try:
#                 import asyncio
#                 loop = asyncio.new_event_loop()
#                 asyncio.set_event_loop(loop)
#                 loop.run_until_complete(send())
#                 loop.close()
#                 print(f"Email sent successfully using alternative method to {recipient_list}")
#             except Exception as e2:
#                 print(f"Alternative email send method also failed: {e2}")
#                 raise e2
        
#         # Clean up files after sending
#         if excel_path and os.path.exists(excel_path):
#             os.remove(excel_path)
#         if pdf_path and os.path.exists(pdf_path):
#             os.remove(pdf_path)
        
#     except Exception as e:
#         print(f"Failed to send production requirements email to {recipient_list if 'recipient_list' in locals() else to}: {e}")
#     finally:
#         db.close()


def send_production(to: Union[str, List[str]], name: str, company_id: int = None, is_update: bool = False, is_email_update: bool = False, recent_order_id: int = None):
    """
    Send production requirements email with attachments
    
    Args:
        to: Recipient email address (string) or list of email addresses
        name: Recipient name
        company_id: Company ID for data filtering
        is_update: Boolean flag to indicate if this is an update to existing requirements
        is_email_update: Boolean flag to indicate if items were ordered after global time
        recent_order_id: ID of the recent order to highlight (when email sent after global time)
    """
    
    def sort_columns_by_location_priority(columns):
        """
        Sort columns to put locations in the specified order
        Returns reordered columns list
        """
        # Define location priority order with all variations
        location_priority = {
            # Midtown East variations - Priority 1
            'midtown east': 1, 'Midtown East': 1,
            # Lenox Hill variations - Priority 2  
            'lenox hill': 2, 'Lenox Hill': 2, 'lenox hills': 2, 'Lenox Hills': 2,
            # Hell's Kitchen variations - Priority 3
            'hells kitchen': 3, 'Hells Kitchen': 3, "hell's kitchen": 3, "Hell's Kitchen": 3,
            # Union Square variations - Priority 4
            'union square': 4, 'Union Square': 4, 'Union square': 4,
            # Flatiron variations - Priority 5
            'flatiron': 5, 'Flatiron': 5, 'FLATIRON': 5,
            # Williamsburg variations - Priority 6
            'williamsburg': 6, 'Williamsburg': 6
        }
        
        def get_column_priority(col):
            """Get sorting priority for a column"""
            col_lower = col.lower()
            
            # Check if it's a location column (contains 'store', 'location', 'shop' or matches known locations)
            is_location_col = any(keyword in col_lower for keyword in ['store', 'location', 'shop']) or col in location_priority
            
            if not is_location_col:
                # Non-location columns get priority 0 (first) or 1000 (last) depending on column type
                if col_lower in ['item', 'item_name', 'name', 'product']:
                    return (0, col)  # Item columns first
                elif any(keyword in col_lower for keyword in ['total', 'required', 'sum']):
                    return (1000, col)  # Total columns last
                else:
                    return (999, col)  # Other columns near the end
            
            # For location columns, check direct match first
            if col in location_priority:
                return (100 + location_priority[col], col)
            
            # Check for partial matches in column name
            for location, priority in location_priority.items():
                if location.lower() in col_lower:
                    return (100 + priority, col)
            
            # Unknown location gets priority after known locations
            return (200, col)
        
        # Sort columns by priority, then alphabetically within same priority
        sorted_columns = sorted(columns, key=get_column_priority)
        return sorted_columns
    
    db = SessionLocal()
    try:
        # Handle both single email and list of emails
        if isinstance(to, str):
            recipient_list = [to]
        elif isinstance(to, list):
            recipient_list = to
        else:
            print(f"Invalid email format: {to}")
            return
        
        print("---------------------------------------------")
        print("Sending production email to:", recipient_list, "for user:", name, "with company ID:", company_id, "Is update:", is_update, "Is email update:", is_email_update, "Recent order ID:", recent_order_id)
        print("---------------------------------------------")
        
        # Get global time for highlighting recent orders
        global_time = None
        recent_order_items = set()
        current_order_items = set()  # Initialize here to ensure it's always available
        
        try:
            first_company_email = db.query(Mail).filter(Mail.company_id == company_id).first()
            if first_company_email and first_company_email.receiving_time:
                global_time = first_company_email.receiving_time
                print(f"Global time for highlighting: {global_time}")
        except Exception as e:
            print(f"Error getting global time: {e}")
        
        data = get_consolidated_production(company_id, db)
        print("Production data retrieved:", bool(data))
        
        if not data:
            print("WARNING: No production data available")
            return  # Exit early if no data
        
        # APPLY COLUMN SORTING HERE - Sort columns by location priority
        if data and 'data' in data and data['data']:
            original_columns = data.get('columns', list(data['data'][0].keys()) if data['data'] else [])
            if original_columns:
                sorted_columns = sort_columns_by_location_priority(original_columns)
                data['columns'] = sorted_columns
                print(f"Columns reordered: {sorted_columns}")
        
        # Get recent orders (after global time) for highlighting
        if global_time:
            try:
                from sqlalchemy import and_
                # Import StoreOrders model - adjust import path as needed
                from models.storeorders import StoreOrders
                
                # Get today's date to combine with global_time
                today = datetime.now().date()
                global_datetime_today = datetime.combine(today, global_time)
                
                # Query orders created after global time
                recent_orders = db.query(StoreOrders).filter(
                    and_(
                        StoreOrders.company_id == company_id,
                        StoreOrders.created_at > global_datetime_today
                    )
                ).all()
                
                # Extract item names from recent orders
                for order in recent_orders:
                    if order.items_ordered and isinstance(order.items_ordered, dict) and 'items' in order.items_ordered:
                        for item in order.items_ordered['items']:
                            if isinstance(item, dict):
                                if 'item_name' in item:
                                    recent_order_items.add(str(item['item_name']))
                                elif 'name' in item:
                                    recent_order_items.add(str(item['name']))
                
                # If we have a specific recent_order_id, get its items for special highlighting
                if recent_order_id:
                    current_order = db.query(StoreOrders).filter(StoreOrders.id == recent_order_id).first()
                    if current_order and current_order.items_ordered and isinstance(current_order.items_ordered, dict) and 'items' in current_order.items_ordered:
                        for item in current_order.items_ordered['items']:
                            if isinstance(item, dict):
                                if 'item_name' in item:
                                    current_order_items.add(str(item['item_name']))
                                elif 'name' in item:
                                    current_order_items.add(str(item['name']))
                        print(f"Current order items to specially highlight: {current_order_items}")
                
                print(f"Recent order items to highlight: {recent_order_items}")
            except Exception as e:
                print(f"Error getting recent orders: {e}")
                # Continue without highlighting if there's an error
        
        print("This is the data which I want to print:", data)

        # Generate timestamp for file names
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create downloads directory if it doesn't exist
        downloads_dir = "downloads"
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)

        # Generate Excel file
        def create_excel_file(data, filename):
            if not data or 'data' not in data or not data['data']:
                return None
            
            try:
                # Use sorted columns for DataFrame creation
                columns = data.get('columns', list(data['data'][0].keys()) if data['data'] else [])
                
                # Create DataFrame with ordered columns
                df_data = []
                for item in data['data']:
                    ordered_row = {col: item.get(col, "") for col in columns}
                    df_data.append(ordered_row)
                
                df = pd.DataFrame(df_data, columns=columns)
                
                # Create Excel file with formatting
                excel_path = os.path.join(downloads_dir, filename)
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Production Requirements', index=False)
                    
                    # Get workbook and worksheet objects
                    workbook = writer.book
                    worksheet = writer.sheets['Production Requirements']
                    
                    # Apply formatting
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    # Header formatting
                    header_font = Font(bold=True, color="000000")
                    header_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                    total_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    highlight_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    recent_order_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red for recent orders
                    current_order_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Darker red for current order
                    
                    # Border style
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Get item name column index
                    item_name_col = None
                    for idx, col in enumerate(df.columns):
                        if 'item' in col.lower() and ('name' in col.lower() or col.lower() == 'item'):
                            item_name_col = idx + 1  # Excel is 1-indexed
                            break
                    
                    # Format headers
                    for col_num, column in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                        
                        # Special formatting for total columns
                        if 'total' in column.lower() or 'required' in column.lower():
                            cell.fill = total_fill
                    
                    # Format data cells
                    for row_num in range(2, len(df) + 2):
                        # Check if this item is a recent order or current order
                        is_recent_item = False
                        is_current_order_item = False
                        if item_name_col and (recent_order_items or current_order_items):
                            item_name = worksheet.cell(row=row_num, column=item_name_col).value
                            item_name_str = str(item_name)
                            is_current_order_item = item_name_str in current_order_items
                            # Recent items include ALL items ordered after global time (including current order)
                            is_recent_item = item_name_str in recent_order_items
                        
                        for col_num, column in enumerate(df.columns, 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            # Apply base effects first: 1) Recent orders (red), 2) Totals (green), 3) Positive values (yellow)
                            if is_recent_item:
                                cell.fill = recent_order_fill
                                cell.font = Font(bold=True, color="CC0000")  # Red bold text
                            elif 'total' in column.lower() or 'required' in column.lower():
                                cell.fill = total_fill
                                cell.font = Font(bold=True)
                            elif ('total' not in column.lower() and 'required' not in column.lower() 
                                and isinstance(cell.value, (int, float)) and cell.value > 0):
                                cell.fill = highlight_fill
                            
                            # Final layer: Apply bold border and underline for current order items (keep existing formatting)
                            if is_current_order_item:
                                # Keep existing font but add underline
                                current_font = cell.font
                                cell.font = Font(
                                    bold=current_font.bold,
                                    color=current_font.color,
                                    italic=False,  # No italic
                                    underline='single'
                                )
                                # Add medium border for current order items
                                medium_border = Border(
                                    left=Side(style='medium'),
                                    right=Side(style='medium'),
                                    top=Side(style='medium'),
                                    bottom=Side(style='medium')
                                )
                                cell.border = medium_border
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                return excel_path
            except Exception as e:
                print(f"Error creating Excel file: {e}")
                return None

        # Generate PDF file
        def create_pdf_file(data, filename):
            if not data or 'data' not in data or not data['data']:
                return None
            
            try:
                pdf_path = os.path.join(downloads_dir, filename)
                
                # Create PDF document
                doc = SimpleDocTemplate(pdf_path, pagesize=A4)
                elements = []
                
                # Get styles
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    spaceAfter=30,
                    alignment=1  # Center alignment
                )
                
                # Add title with update indicator if needed
                title_text = "Consolidated Production Requirements"
                if is_update:
                    title_text += " (Updated)"
                title = Paragraph(title_text, title_style)
                elements.append(title)
                elements.append(Spacer(1, 12))
                
                # Add subtitle
                subtitle = Paragraph("", styles['Normal'])
                elements.append(subtitle)
                elements.append(Spacer(1, 20))
                
                # Prepare table data with sorted columns
                columns = data.get('columns', list(data['data'][0].keys()) if data['data'] else [])
                table_data = [columns]  # Header row
                
                # Add data rows with ordered columns
                for item in data['data']:
                    row = [str(item.get(col, "")) for col in columns]
                    table_data.append(row)
                
                # Create table
                table = Table(table_data)
                
                # Define table style
                table_style = [
                    # Header row styling
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    
                    # Data rows styling
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]
                
                # Highlight total columns and positive values
                for col_idx, col in enumerate(columns):
                    if 'total' in col.lower() or 'required' in col.lower():
                        # Green background for total columns
                        table_style.append(('BACKGROUND', (col_idx, 0), (col_idx, -1), colors.lightgreen))
                        table_style.append(('FONTNAME', (col_idx, 1), (col_idx, -1), 'Helvetica-Bold'))
                
                # Apply alternating row colors
                for row_idx in range(1, len(table_data)):
                    if row_idx % 2 == 0:
                        table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.beige))
                
                table.setStyle(TableStyle(table_style))
                elements.append(table)
                
                # Add footer
                elements.append(Spacer(1, 30))
                footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>KPI360.ai Team"
                footer = Paragraph(footer_text, styles['Normal'])
                elements.append(footer)
                
                # Build PDF
                doc.build(elements)
                return pdf_path
            except Exception as e:
                print(f"Error creating PDF file: {e}")
                return None

        # Generate HTML table from the data - DYNAMIC VERSION with sorted columns
        def generate_production_table(data):
            if not data or 'data' not in data or not data['data']:
                return "<p>No production data available.</p>"
            
            # Get sorted columns from the data
            columns = data.get('columns', [])
            if not columns and data['data']:
                # If columns not provided, extract from first row and sort them
                columns = sort_columns_by_location_priority(list(data['data'][0].keys()))
            
            if not columns:
                return "<p>No columns found in data.</p>"
            
            # Find item name column
            item_name_col = None
            for idx, col in enumerate(columns):
                if 'item' in col.lower() and ('name' in col.lower() or col.lower() == 'item'):
                    item_name_col = idx
                    break
            
            # Start building the table
            table_html = """
            <table style="border-collapse: collapse; width: 100%; margin: 20px 0;">
                <thead>
                    <tr style="background-color: #f8f9fa;">
            """
            
            # Generate header row dynamically with sorted columns
            for col in columns:
                # Special styling for specific column types
                if 'total' in col.lower() or 'required' in col.lower() or 'sum' in col.lower():
                    header_style = "border: 1px solid #ddd; padding: 12px; text-align: center; font-weight: bold; background-color: #e8f5e8;"
                else:
                    header_style = "border: 1px solid #ddd; padding: 12px; text-align: center; font-weight: bold;"
                
                table_html += f'<th style="{header_style}">{col}</th>'
            
            table_html += """
                    </tr>
                </thead>
                <tbody>
            """
            
            # Generate data rows dynamically with sorted columns
            for row_index, item in enumerate(data['data']):
                # Check if this item is a recent order or current order
                is_recent_item = False
                is_current_order_item = False
                if item_name_col is not None and (recent_order_items or current_order_items):
                    item_name = list(item.get(col, "") for col in columns)[item_name_col] if item_name_col < len(columns) else None
                    item_name_str = str(item_name)
                    is_current_order_item = item_name_str in current_order_items
                    # Recent items include ALL items ordered after global time (including current order)
                    is_recent_item = item_name_str in recent_order_items
                
                # Alternate row colors for better readability, with special highlighting for orders
                if is_recent_item:
                    row_bg = "#ffe6e6"  # Light red for recent orders
                else:
                    row_bg = "#f9f9f9" if row_index % 2 == 0 else "#ffffff"
                
                table_html += f'<tr style="background-color: {row_bg};">'
                
                for col_index, col in enumerate(columns):
                    cell_value = item.get(col, "")
                    
                    # Apply base effects first: 1) Recent orders (red), 2) Totals (green), 3) Positive values (yellow)
                    if is_recent_item and col_index == item_name_col:
                        # Recent item name - bold red text
                        cell_style = "border: 1px solid #ddd; padding: 12px; text-align: left; background-color: #ffe6e6; font-weight: bold; color: #cc0000;"
                    elif is_recent_item:
                        # Recent item other columns
                        cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center; background-color: #ffe6e6; font-weight: bold; color: #cc0000;"
                    elif 'total' in col.lower() or 'required' in col.lower() or 'sum' in col.lower():
                        # Highlight total/required columns in green
                        cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center; background-color: #e8f5e8; font-weight: bold;"
                    elif col_index == 0:  # First column (usually item name)
                        # Check if any store column has a value for this row to highlight the item
                        has_store_value = any(
                            isinstance(item.get(c, 0), (int, float)) and item.get(c, 0) > 0 
                            for c in columns[1:-2] if 'store' in c.lower() or 'location' in c.lower() or 'shop' in c.lower()
                        )
                        if has_store_value:
                            cell_style = "border: 1px solid #ddd; padding: 12px; text-align: left; background-color: #fff3cd;"
                        else:
                            cell_style = "border: 1px solid #ddd; padding: 12px; text-align: left;"
                    elif isinstance(cell_value, (int, float)) and cell_value > 0:
                        # Highlight cells with positive values (like store quantities)
                        cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center; background-color: #fff3cd;"
                    elif col.lower() in ['unit', 'units', 'type', 'category']:
                        # Unit/type columns - center aligned, no highlighting
                        cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center;"
                    else:
                        # Default styling
                        cell_style = "border: 1px solid #ddd; padding: 12px; text-align: center;"
                    
                    # Final layer: Apply medium border and underline for current order items (keep base effects)
                    if is_current_order_item:
                        # Add underline and medium border to existing styling
                        cell_style += " text-decoration: underline; border: 2px solid #333;"
                    
                    table_html += f'<td style="{cell_style}">{cell_value}</td>'
                
                table_html += "</tr>"
            
            table_html += """
                </tbody>
            </table>
            """
            
            return table_html

        # Generate files
        excel_filename = f"production_requirements_{timestamp}.xlsx"
        pdf_filename = f"production_requirements_{timestamp}.pdf"
        
        excel_path = create_excel_file(data, excel_filename)
        pdf_path = create_pdf_file(data, pdf_filename)
        
        production_table = generate_production_table(data)
        
        # Determine subject and greeting based on flags
        subject = "Consolidated Production Requirements"
        greeting = f"Hello {name},"
        
        # Determine intro text based on is_email_update flag
        if is_email_update:
            intro_text = "**Updated** - Please find below the *updated* consolidated production requirements for all stores."
        else:
            intro_text = "Please find below the consolidated production requirements for all stores."
        
        # Add "Updated" to subject if is_update is True
        if is_update:
            subject += " - Updated"
        
        html_body = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    max-width: 1200px;
                    margin: 0 auto;
                    padding: 20px;
                }}
                .header {{
                    background-color: #f8f9fa;
                    padding: 20px;
                    border-radius: 5px;
                    margin-bottom: 20px;
                    text-align: center;
                }}
                .footer {{
                    margin-top: 30px;
                    padding-top: 20px;
                    border-top: 1px solid #ddd;
                }}
                .table-container {{
                    overflow-x: auto;
                }}
                @media screen and (max-width: 600px) {{
                    .table-container {{
                        font-size: 12px;
                    }}
                }}
                .update-notice {{
                    background-color: #fff3cd;
                    border: 1px solid #ffeaa7;
                    border-radius: 5px;
                    padding: 15px;
                    margin-bottom: 20px;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>{subject}</h2>
            </div>
            
            <h3>{greeting}</h3>
            
            {"<div class='update-notice'><strong>📋 Update Notice:</strong> This is an updated version of the production requirements.</div>" if is_update else ""}
            
            {"<div class='update-notice'><strong>🔔 New Orders Alert:</strong> Some items were ordered after the global time and are highlighted in red." + (f" Items from Order ID #{recent_order_id} also have medium borders and underlines." if recent_order_id else "") + "</div>" if is_email_update else ""}
            
            <p>{intro_text} Excel and PDF versions are attached to this email for your convenience.</p>
            
            <div class="table-container">
                {production_table}
            </div>
            
            <div class="footer">
                <p>This report shows the total quantities needed for production across all stores.</p>
                <p><strong>Legend:</strong></p>
                <ul>
                    <li>🔴 <strong>Red highlighted rows:</strong> Items ordered after global time (recent orders)</li>
                    <li>🟡 <strong>Yellow highlighted cells:</strong> Items required by specific stores</li>
                    <li>🟢 <strong>Green highlighted columns:</strong> Total/Required quantities</li>
                    <li>📊 <strong>Alternating row colors:</strong> For better readability</li>
                    {f"<li>🔲 <strong><u>Medium borders and underlined:</u></strong> Items from current order (Order #{recent_order_id})</li>" if recent_order_id else ""}
                </ul>
                <p>This report automatically adapts to your data structure and orders locations by priority: Midtown East, Lenox Hill, Hell's Kitchen, Union Square, Flatiron, Williamsburg, then others.</p>
                
                <p>Regards,<br><strong>KPI360.ai Team</strong></p>
            </div>
        </body>
        </html>
        """

        # Create attachments list for your MessageSchema
        attachments = []
        
        # Add Excel attachment
        if excel_path and os.path.exists(excel_path):
            attachments.append(excel_path)
        
        # Add PDF attachment  
        if pdf_path and os.path.exists(pdf_path):
            attachments.append(pdf_path)

        # Use your original MessageSchema with attachments
        message = MessageSchema(
            subject=subject,
            recipients=recipient_list,  # Now supports multiple recipients
            body=html_body,
            subtype="html",
            attachments=attachments if attachments else None  # Add attachments if they exist
        )

        async def send():
            try:
                await fm.send_message(message)
                print(f"Email sent successfully to {recipient_list}")
            except Exception as e:
                print(f"Error sending email: {e}")
                raise e

        try:
            asyncio.run(send())
            print(f"Production requirements email {'(update)' if is_update else ''} {'(with recent orders)' if is_email_update else ''} with downloads sent successfully to {recipient_list}")
        except Exception as e:
            print(f"Failed to send email via asyncio: {e}")
            # Try alternative approach if asyncio.run fails
            try:
                import asyncio
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                loop.run_until_complete(send())
                loop.close()
                print(f"Email sent successfully using alternative method to {recipient_list}")
            except Exception as e2:
                print(f"Alternative email send method also failed: {e2}")
                raise e2
        
        # Clean up files after sending
        if excel_path and os.path.exists(excel_path):
            os.remove(excel_path)
        if pdf_path and os.path.exists(pdf_path):
            os.remove(pdf_path)
        
    except Exception as e:
        print(f"Failed to send production requirements email to {recipient_list if 'recipient_list' in locals() else to}: {e}")
    finally:
        db.close()



